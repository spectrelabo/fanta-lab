#!/usr/bin/env python3
"""
Generatore File Excel Ufficiale in ITALIANO per la Lega e per gli Amici.

Genera un file Excel multi-foglio perfettamente formattato in italiano:
  - TUTTI I GIOCATORI (Listone integrale con score composito e prezzo fair)
  - PORTIERI (Con Clean Sheet %, affidabilità e gerarchia)
  - DIFENSORI (Con voti da modificatore e propensione bonus)
  - CENTROCAMPISTI (Con metriche xG/xA e probabilità bonus)
  - ATTACCANTI (Con proiezioni punti P10/P50/P90 e fasce d'asta)
  - LEGENDA METRICHE (Guida dettagliata in italiano a tutte le colonne)

Output:
  - data/analisi_fantacalcio_italiano.xlsx
"""

import os, sys, warnings
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import pandas as pd
import numpy as np

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
import config

warnings.filterwarnings("ignore")

OUTPUT_EXCEL_IT = os.path.join(config.DATA_DIR, "analisi_fantacalcio_italiano.xlsx")

DISPLAY_COLS = [
    "player", "role", "role_mantra", "team", "Prezzo_Consigliato_Cr",
    "prezzo_fair_1000", "surplus_value_cr", "score_composito",
    "predicted_pts_p50", "predicted_pts_p10", "predicted_pts_p90", "pts_volatility_spread",
    "vorp_points", "mv_media_3y", "mv_std", "mv_trend", "availability",
    "xg_media_3y", "xa_media_3y", "offensive_index",
    "giorni_infortunio_3y", "n_infortuni_3y", "infortunio_grave", "malus_infortuni",
    "NN_MV_Atteso", "NN_FV_Atteso", "Prob_Bonus_Ge_8_%", "Prob_Picco_Ge_10_%",
    "Clean_Sheet_%", "FVM_1000"
]

COL_RENAME_IT = {
    "player": "Calciatore",
    "role": "Ruolo",
    "role_mantra": "Ruolo Mantra",
    "team": "Squadra",
    "Prezzo_Consigliato_Cr": "Prezzo Listone (Cr)",
    "prezzo_fair_1000": "Prezzo Fair Asta (su 1000)",
    "surplus_value_cr": "Surplus Valore (Cr)",
    "score_composito": "Score Sintetico (0-1)",
    "predicted_pts_p50": "Punti Attesi (P50)",
    "predicted_pts_p10": "Punti Minimi Floor (P10)",
    "predicted_pts_p90": "Punti Picco Ceiling (P90)",
    "pts_volatility_spread": "Forbice Volatilità Punti",
    "vorp_points": "Punti VORP",
    "mv_media_3y": "Media Voto Storica (3y)",
    "mv_std": "Volatilità Voto (Std)",
    "mv_trend": "Trend Rendimento",
    "availability": "Presenze Disponibilità %",
    "xg_media_3y": "xG Medi (3y)",
    "xa_media_3y": "xA Medi (3y)",
    "offensive_index": "Indice Offensivo Squadra",
    "giorni_infortunio_3y": "Giorni Infortunio (3y)",
    "n_infortuni_3y": "N. Infortuni (3y)",
    "infortunio_grave": "Flag Infortunio Grave",
    "malus_infortuni": "Malus Infortuni",
    "NN_MV_Atteso": "Media Voto Base Attesa",
    "NN_FV_Atteso": "FantaMedia Base Attesa",
    "Prob_Bonus_Ge_8_%": "Prob. Bonus (>=8) %",
    "Prob_Picco_Ge_10_%": "Prob. Picco (>=10) %",
    "Clean_Sheet_%": "Clean Sheet %",
    "FVM_1000": "FVM Ufficiale (su 1000)"
}

LEGENDA_IT = [
    ("Calciatore / Ruolo / Squadra", "Nome ufficiale del giocatore, ruolo classico (P, D, C, A), ruoli Mantra e squadra di appartenenza a mercato chiuso."),
    ("Prezzo Listone (Cr)", "Quotazione iniziale ufficiale pubblicata da Fantacalcio.it."),
    ("Prezzo Fair Asta (su 1000)", "Valutazione economica razionale calibrata sull'asta a 10 partecipanti (1000 crediti), basata su scarsità di reparto, modificatore difesa e proiezioni gol/assist."),
    ("Surplus Valore (Cr)", "Differenza tra Prezzo Fair e Prezzo di Listone. Valori positivi indicano occasioni sottovalutate dal mercato; valori negativi segnalano giocatori 'trappola' o sopravvalutati."),
    ("Score Sintetico (0-1)", "Indice composito globale normalizzato (0.00 - 1.00) che pesa voti puri, expected goals/assists, continuità di rendimento e tenuta fisica."),
    ("Punti Attesi (P50)", "Punteggio totale atteso a fine campionato calcolato con regressione quantilica a Gradient Boosting (mediana statistica)."),
    ("Punti Minimi Floor (P10)", "Scenario pessimistico (10° percentile) in caso di stagione difficile, ballottaggi o lievi stop."),
    ("Punti Picco Ceiling (P90)", "Scenario ottimistico (90° percentile) in caso di exploit realizzativo o stagione d'oro."),
    ("Forbice Volatilità Punti", "Differenza tra Ceiling (P90) e Floor (P10). Forbice stretta = certezza di rendimento; forbice larga = scommessa ad alto rischio/rendimento."),
    ("Punti VORP", "Value Over Replacement Player: punti marginali generati rispetto all'ultimo calciatore titolare prendibile a 1 credito sul mercato degli svincolati."),
    ("Media Voto Storica (3y)", "Media voto pura delle ultime 3 stagioni ponderata per freschezza temporale (più peso alla stagione recente)."),
    ("Volatilità Voto (Std)", "Deviazione standard dei voti. Più è bassa, più il giocatore è costante e sicuro per il modificatore."),
    ("Trend Rendimento", "Pendenza della crescita o del calo del rendimento nelle ultime tre stagioni (+ in ascesa, - in calo)."),
    ("Presenze Disponibilità %", "Percentuale di partite disputate negli ultimi 3 anni rispetto alle 38 disponibili."),
    ("xG Medi (3y) / xA Medi (3y)", "Expected Goals ed Expected Assists per 90 minuti ricavati dai modelli di tiro e passaggio Understat."),
    ("Indice Offensivo Squadra", "Forza offensiva stimata della squadra di appartenenza (capacità di creare occasioni da gol)."),
    ("Giorni / N. Infortuni (3y)", "Storico dei giorni saltati per infortunio e numero totale di stop registrati da Transfermarkt negli ultimi 3 anni."),
    ("Flag Infortunio Grave", "Segnala se il calciatore ha subito lesioni importanti (es. crociato, menisco, tendini) nell'ultimo triennio."),
    ("Clean Sheet %", "Percentuale storica di partite concluse a porta inviolata (cruciale per portieri e difensori da imbattibilità)."),
    ("FVM Ufficiale (su 1000)", "FantaValore di Mercato ufficiale di riferimento su scala 1.000 crediti.")
]


def create_italian_workbook():
    print("=" * 60)
    print("  GENERAZIONE FILE EXCEL IN ITALIANO (per amici e lega)")
    print("=" * 60)

    if not os.path.exists(config.DATASET_FINALE_CSV):
        raise FileNotFoundError(f"File {config.DATASET_FINALE_CSV} non trovato.")

    df = pd.read_csv(config.DATASET_FINALE_CSV)

    # Assicura la presenza di tutte le colonne
    for c in DISPLAY_COLS:
        if c not in df.columns:
            df[c] = np.nan

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Rimuove il foglio vuoto iniziale

    # Stili grafici professionali
    font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    fill_header = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")  # Navy Blue
    fill_header_legenda = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")

    fill_zebra = PatternFill(start_color="F2F5F9", end_color="F2F5F9", fill_type="solid")
    fill_white = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    font_data = Font(name="Calibri", size=10)
    font_bold = Font(name="Calibri", size=10, bold=True)
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")

    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9")
    )

    sheets_to_create = [
        ("TUTTI I GIOCATORI", df.sort_values("prezzo_fair_1000", ascending=False)),
        ("PORTIERI", df[df["role"] == "P"].sort_values("prezzo_fair_1000", ascending=False)),
        ("DIFENSORI", df[df["role"] == "D"].sort_values("prezzo_fair_1000", ascending=False)),
        ("CENTROCAMPISTI", df[df["role"] == "C"].sort_values("prezzo_fair_1000", ascending=False)),
        ("ATTACCANTI", df[df["role"] == "A"].sort_values("prezzo_fair_1000", ascending=False))
    ]

    for title, sub_df in sheets_to_create:
        ws = wb.create_sheet(title=title)
        ws.views.sheetView[0].showGridLines = True

        # Intestazioni in italiano
        headers_it = [COL_RENAME_IT.get(c, c) for c in DISPLAY_COLS]
        ws.append(headers_it)

        # Stile Header
        for col_idx in range(1, len(headers_it) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border
        ws.row_dimensions[1].height = 28

        # Righe Dati
        for row_i, (_, row_data) in enumerate(sub_df.iterrows(), start=2):
            row_vals = []
            for col_key in DISPLAY_COLS:
                val = row_data.get(col_key)
                if pd.isna(val):
                    row_vals.append("")
                elif col_key in ["prezzo_fair_1000", "Prezzo_Consigliato_Cr", "surplus_value_cr", "FVM_1000", "giorni_infortunio_3y", "n_infortuni_3y"]:
                    row_vals.append(int(round(float(val))))
                elif col_key in ["score_composito", "predicted_pts_p50", "predicted_pts_p10", "predicted_pts_p90", "pts_volatility_spread", "vorp_points", "mv_media_3y", "mv_std", "xg_media_3y", "xa_media_3y", "NN_MV_Atteso", "NN_FV_Atteso"]:
                    row_vals.append(round(float(val), 2))
                elif col_key in ["availability", "Prob_Bonus_Ge_8_%", "Prob_Picco_Ge_10_%", "Clean_Sheet_%"]:
                    row_vals.append(f"{round(float(val), 1)}%")
                else:
                    row_vals.append(val)

            ws.append(row_vals)
            row_fill = fill_zebra if row_i % 2 == 0 else fill_white

            for col_idx, col_key in enumerate(DISPLAY_COLS, start=1):
                c_cell = ws.cell(row=row_i, column=col_idx)
                c_cell.font = font_bold if col_key in ["player", "prezzo_fair_1000", "score_composito"] else font_data
                c_cell.fill = row_fill
                c_cell.border = thin_border

                if col_key in ["player", "team", "role_mantra"]:
                    c_cell.alignment = align_left
                elif col_key in ["role", "infortunio_grave"]:
                    c_cell.alignment = align_center
                else:
                    c_cell.alignment = align_right

            ws.row_dimensions[row_i].height = 20

        # Auto-fit larghezza colonne
        for col_idx, col_name in enumerate(headers_it, start=1):
            col_letter = get_column_letter(col_idx)
            max_len = max(len(str(col_name)), max((len(str(ws.cell(r, col_idx).value or '')) for r in range(2, min(len(sub_df) + 2, 50))), default=5))
            ws.column_dimensions[col_letter].width = min(max(max_len + 3, 11), 32)

        # Attiva i filtri automatici su tutte le colonne
        ws.auto_filter.ref = f"A1:{get_column_letter(len(DISPLAY_COLS))}{len(sub_df) + 1}"
        ws.freeze_panes = "B2"  # Blocca la prima colonna giocatore e l'header

    # Foglio LEGENDA METRICHE in Italiano
    ws_leg = wb.create_sheet(title="LEGENDA METRICHE")
    ws_leg.views.sheetView[0].showGridLines = True
    ws_leg.append(["Metrica / Colonna", "Spiegazione Dettagliata & Utilizzo Pratico all'Asta"])

    for col_idx in [1, 2]:
        cell = ws_leg.cell(row=1, column=col_idx)
        cell.font = font_header
        cell.fill = fill_header_legenda
        cell.alignment = align_left
        cell.border = thin_border
    ws_leg.row_dimensions[1].height = 26

    for r_idx, (m_title, m_desc) in enumerate(LEGENDA_IT, start=2):
        ws_leg.append([m_title, m_desc])
        row_fill = fill_zebra if r_idx % 2 == 0 else fill_white

        c1 = ws_leg.cell(row=r_idx, column=1)
        c1.font = font_bold
        c1.fill = row_fill
        c1.alignment = align_left
        c1.border = thin_border

        c2 = ws_leg.cell(row=r_idx, column=2)
        c2.font = font_data
        c2.fill = row_fill
        c2.alignment = align_left
        c2.border = thin_border
        ws_leg.row_dimensions[r_idx].height = 24

    ws_leg.column_dimensions["A"].width = 30
    ws_leg.column_dimensions["B"].width = 110

    wb.save(OUTPUT_EXCEL_IT)
    print(f"\n  ✅ File Excel Ufficiale in Italiano salvato: {OUTPUT_EXCEL_IT}")
    print("  STAGE ITALIAN EXCEL COMPLETATO.\n")


if __name__ == "__main__":
    create_italian_workbook()
