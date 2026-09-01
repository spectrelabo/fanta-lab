#!/usr/bin/env python3
"""
FASE 5 — Generazione Excel multi-scheda formattato.

Crea un file Excel professionale con:
  - LISTONE COMPLETO (tutti i giocatori attivi ordinati per score)
  - PORTIERI / DIFENSORI / CENTROCAMPISTI / ATTACCANTI (schede per ruolo)
  - LEGENDA COLONNE (guida al significato di ogni metrica)

Output:
  - analisi_fantacalcio_completa.xlsx
"""

import os, sys, warnings
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import pandas as pd
import numpy as np

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
import config

warnings.filterwarnings("ignore")

# Colonne da mostrare nell'Excel (ordine logico)
DISPLAY_COLS = [
    "player", "role", "role_mantra", "team", "Prezzo_Consigliato_Cr", "score_composito",
    "mv_media_3y", "mv_std", "mv_trend", "availability",
    "xg_media_3y", "xa_media_3y", "offensive_index",
    "giorni_infortunio_3y", "n_infortuni_3y", "infortunio_grave", "malus_infortuni",
    "NN_MV_Atteso", "NN_FV_Atteso", "Prob_Bonus_Ge_8_%", "Prob_Picco_Ge_10_%",
    "Clean_Sheet_%", "FVM_1000"
]

# Rinomina per intestazioni Excel leggibili
COL_RENAME = {
    "player": "Giocatore", "role": "Ruolo", "role_mantra": "Ruolo Mantra",
    "team": "Squadra", "Prezzo_Consigliato_Cr": "Prezzo Asta (Cr)",
    "score_composito": "Score Finale",
    "mv_media_3y": "MV Ponderata (3y)", "mv_std": "Volatilità MV (Std)",
    "mv_trend": "Trend MV", "availability": "Disponibilità %",
    "xg_media_3y": "xG Medio (3y)", "xa_media_3y": "xA Medio (3y)",
    "offensive_index": "Indice Off. Squadra",
    "giorni_infortunio_3y": "Giorni Stop (3y)",
    "n_infortuni_3y": "N. Infortuni (3y)",
    "infortunio_grave": "Infortunio Grave", "malus_infortuni": "Malus Infortuni",
    "NN_MV_Atteso": "Stima MV", "NN_FV_Atteso": "Stima FM",
    "Prob_Bonus_Ge_8_%": "Prob. Bonus ≥8 %",
    "Prob_Picco_Ge_10_%": "Prob. Picco ≥10 %",
    "Clean_Sheet_%": "Clean Sheet %", "FVM_1000": "FVM (su 1000)"
}

# Legenda colonne per la guida all'asta
LEGENDA = [
    ("Score Finale", "Punteggio sintetico globale (0.00-1.00). Combina MV storica, stime rendimento, xG, regolarità presenze e affidabilità fisica."),
    ("Prezzo Asta (Cr)", "Quotazione ufficiale Fantacalcio.it. Riferimento per il rilancio massimo all'asta (su budget 1000 crediti)."),
    ("MV Ponderata (3y)", "Media voto ultimi 3 anni con pesi decrescenti (3× recente, 2× precedente, 1× meno recente). Riduce dipendenza da singola annata."),
    ("Volatilità MV (Std)", "Deviazione standard MV. Valori bassi (<0.20) = rendimento costante; alti (>0.50) = forte altalenanza."),
    ("Trend MV", "Inclinazione tendenza rendimento. Positivi = in crescita; Negativi = in calo."),
    ("Disponibilità %", "% partite a voto su 38 giornate teoriche (media anni attivi in Serie A)."),
    ("xG Medio (3y)", "Expected Goals medi per stagione calcolati da Understat."),
    ("xA Medio (3y)", "Expected Assists medi per stagione (qualità dei passaggi chiave)."),
    ("Indice Off. Squadra", "Fattore correttivo potenza offensiva squadra (range 0.90-1.10)."),
    ("Giorni Stop (3y)", "Totale giorni assenza per infortunio ultimi 3 anni (Transfermarkt)."),
    ("N. Infortuni (3y)", "Numero totale eventi stop ultimi 3 anni."),
    ("Infortunio Grave", "1=Sì, 0=No — almeno un infortunio >60 giorni continui."),
    ("Malus Infortuni", "Indice fragilità fisica (0.0=Integro, 1.0=Altissima fragilità). Penalità fino -15% sullo score."),
    ("Stima MV", "Media voto attesa dal modello Fantacalcio.it per la stagione corrente."),
    ("Stima FM", "Fantamedia attesa (voto + bonus/malus) dal modello per la stagione corrente."),
    ("Prob. Bonus ≥8 %", "Probabilità stimata di punteggio ≥8.0 in una singola giornata."),
    ("Prob. Picco ≥10 %", "Probabilità di prestazione da picco assoluto (≥10.0)."),
    ("Clean Sheet %", "% partite senza subire reti (fondamentale per portieri e difensori)."),
    ("FVM (su 1000)", "Fantavalore di Mercato ufficiale espresso in millesimi.")
]


def generate_excel(df, output_path=None):
    """Genera il workbook Excel multi-scheda formattato."""
    output_path = output_path or config.OUTPUT_EXCEL
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Stili
    header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    role_fills = {
        "P": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
        "D": PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
        "C": PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid"),
        "A": PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),
    }
    border_thin = Border(
        left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9')
    )
    align_center = Alignment(horizontal="center", vertical="center")
    align_left   = Alignment(horizontal="left", vertical="center")
    align_right  = Alignment(horizontal="right", vertical="center")

    available_cols = [c for c in DISPLAY_COLS if c in df.columns]

    def format_sheet(ws, df_subset):
        df_sub = df_subset[available_cols].rename(columns=COL_RENAME)
        headers = list(df_sub.columns)
        ws.append(headers)

        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = align_center
            cell.border = border_thin

        for r_idx, row in enumerate(df_sub.itertuples(index=False), start=2):
            role_val = row[1]
            fill_bg = role_fills.get(role_val, PatternFill(fill_type=None))
            for c_idx, val in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx)
                if pd.isna(val) or val is None:
                    cell.value = "-"
                    cell.alignment = align_center
                elif isinstance(val, (float, np.floating)):
                    col_name = headers[c_idx - 1]
                    if "%" in col_name or "Disponibilità" in col_name:
                        cell.value = val
                        cell.number_format = '0.0%' if val <= 1.0 else '0.0'
                    elif "Score" in col_name or "Trend" in col_name or "Malus" in col_name:
                        cell.value = round(val, 4)
                        cell.number_format = '0.0000'
                    else:
                        cell.value = round(val, 2)
                        cell.number_format = '0.00'
                    cell.alignment = align_right
                elif isinstance(val, (int, np.integer)):
                    cell.value = int(val)
                    cell.alignment = align_right
                else:
                    cell.value = str(val)
                    cell.alignment = align_left if c_idx in [1, 4] else align_center
                if c_idx == 2:
                    cell.fill = fill_bg
                cell.border = border_thin

        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    # Crea schede
    format_sheet(wb.create_sheet(title="LISTONE COMPLETO"), df)
    format_sheet(wb.create_sheet(title="PORTIERI"), df[df["role"] == "P"])
    format_sheet(wb.create_sheet(title="DIFENSORI"), df[df["role"] == "D"])
    format_sheet(wb.create_sheet(title="CENTROCAMPISTI"), df[df["role"] == "C"])
    format_sheet(wb.create_sheet(title="ATTACCANTI"), df[df["role"] == "A"])

    # Legenda
    ws_leg = wb.create_sheet(title="LEGENDA COLONNE")
    ws_leg.append(["Colonna", "Descrizione e Significato Operativo all'Asta"])
    ws_leg.cell(row=1, column=1).fill = header_fill
    ws_leg.cell(row=1, column=1).font = header_font
    ws_leg.cell(row=1, column=2).fill = header_fill
    ws_leg.cell(row=1, column=2).font = header_font
    for r_idx, (col_name, desc) in enumerate(LEGENDA, start=2):
        c1 = ws_leg.cell(row=r_idx, column=1, value=col_name)
        c2 = ws_leg.cell(row=r_idx, column=2, value=desc)
        c1.font = Font(name="Calibri", size=11, bold=True)
        c1.alignment = align_left
        c2.alignment = align_left
        c1.border = border_thin
        c2.border = border_thin
    ws_leg.column_dimensions['A'].width = 25
    ws_leg.column_dimensions['B'].width = 110

    wb.save(output_path)
    print(f"  ✅ Excel salvato: {output_path}")


def main():
    """Esegue Fase 5: generazione Excel."""
    print("=" * 60)
    print("  FASE 5 — GENERAZIONE EXCEL MULTI-SCHEDA")
    print("=" * 60)

    df = pd.read_csv(config.DATASET_FINALE_CSV)
    print(f"\n  Giocatori nel dataset: {len(df)}")

    generate_excel(df)

    for role, name in [("P", "Portieri"), ("D", "Difensori"), ("C", "Centrocampisti"), ("A", "Attaccanti")]:
        count = len(df[df["role"] == role])
        print(f"  {name}: {count}")

    print("\n  FASE 5 COMPLETATA.\n")


if __name__ == "__main__":
    main()
